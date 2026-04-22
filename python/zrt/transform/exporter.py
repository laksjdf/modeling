"""Export transformed OpGraph to Excel, JSON, and ONNX with parallelism annotations."""
from __future__ import annotations

import json
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from python.zrt.ir.graph import OpGraph
from python.zrt.ir.node import OpNode
from python.zrt.transform.context import TransformContext, ParallelConfig

logger = logging.getLogger(__name__)


def infer_pipeline_stage(node: OpNode, layer_to_stage: Optional[Dict[str, int]] = None) -> str:
    """Infer pipeline stage from layer index.

    In pipeline parallelism, consecutive layers are grouped into stages.
    For now, if layer info is available, use layer number; otherwise infer from scope.
    """
    if layer_to_stage and node.layer in layer_to_stage:
        return f"stage_{layer_to_stage[node.layer]}"

    # Try to extract layer number from scope (e.g., "layers.0.mlp" → layer "0")
    if node.layer:
        try:
            stage_num = int(node.layer) // 4  # assume 4 layers per stage as default
            return f"stage_{stage_num}"
        except (ValueError, TypeError):
            pass

    return "stage_0"


def format_stream_info(node: OpNode) -> str:
    """Format stream assignment info as readable string."""
    stream_id = node.annotations.get("stream_id")
    stream_type = node.annotations.get("stream_type")

    if stream_id is None:
        return ""

    if stream_type == "comm":
        return f"comm_stream_{stream_id}"
    else:
        return f"compute_stream_{stream_id}"


def get_parallelism_info(node: OpNode, parallel_config: ParallelConfig) -> Dict[str, str]:
    """Extract parallelism information from node annotations and attributes."""
    result = {
        "strategy": parallel_config.describe(),
        "collective": "",
        "group_size": "",
        "role": "",
    }

    # Check if this is a communication node
    if node.is_comm:
        result["collective"] = node.attrs.get("collective", "")
        result["group_size"] = str(node.attrs.get("group_size", ""))
        result["role"] = node.attrs.get("role", "")

        # Infer which parallel dimension this comm belongs to
        collective = result["collective"]
        group_size = node.attrs.get("group_size", 1)

        if collective == "all_reduce":
            result["parallel_type"] = "TP"  # Tensor Parallel
        elif collective == "all_to_all":
            result["parallel_type"] = "EP"  # Expert Parallel
        else:
            result["parallel_type"] = ""
    else:
        # Check if node has parallel annotations from Split passes
        tp_split = node.annotations.get("tp_split", {})
        ep_annot = node.annotations.get("ep_needs_a2a")

        parallel_types = []
        if tp_split:
            parallel_types.append("TP")
        if ep_annot:
            parallel_types.append("EP")

        result["parallel_type"] = "/".join(parallel_types) or ""

    return result


class TransformedGraphExcelWriter:
    """Write transformed OpGraph to Excel with parallelism, communication, and stream info."""

    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        self._header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        self._header_font = Font(bold=True, color="FFFFFF", size=11)
        self._comm_fill = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
        self._compute_fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
        self._memory_fill = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
        self._thin_border = Border(bottom=Side(style="thin", color="BDBDBD"))

    def write(self, graph: OpGraph, ctx: TransformContext, output_path: Path) -> None:
        """Write transformed graph to Excel with all annotations."""
        wb = openpyxl.Workbook()

        self._write_metadata_sheet(wb, graph, ctx)
        self._write_transformed_ops_sheet(wb, graph, ctx)
        self._write_communication_sheet(wb, graph, ctx)
        self._write_parallelism_summary_sheet(wb, graph, ctx)
        self._write_stream_assignment_sheet(wb, graph, ctx)

        wb.save(output_path)
        logger.info(f"Exported transformed graph to {output_path}")

    def _write_metadata_sheet(self, wb: openpyxl.Workbook,
                              graph: OpGraph, ctx: TransformContext) -> None:
        """Write graph metadata and configuration."""
        ws = wb.active
        ws.title = "Metadata"

        ws.append(["Graph Metadata"])
        ws["A1"].font = Font(bold=True, size=12)

        metadata = [
            ("Graph Name", graph.name),
            ("Phase", graph.phase),
            ("Total Nodes", len(graph.nodes)),
            ("Total Edges", len(graph.edges)),
            ("", ""),
            ("Parallelism Config", ""),
            ("  TP (Tensor Parallel)", ctx.parallel.tp),
            ("  EP (Expert Parallel)", ctx.parallel.ep),
            ("  PP (Pipeline Parallel)", ctx.parallel.pp),
            ("  DP (Data Parallel)", ctx.parallel.dp),
            ("  Sequence Parallel", "Yes" if ctx.parallel.sp else "No"),
            ("  Strategy Description", ctx.parallel.describe()),
            ("", ""),
            ("Stream Config", ""),
            ("  Compute Streams", ctx.stream_config.num_compute_streams),
            ("  Comm Streams", ctx.stream_config.num_comm_streams),
        ]

        for key, value in metadata:
            ws.append([key, value])

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = self._thin_border

    def _write_transformed_ops_sheet(self, wb: openpyxl.Workbook,
                                     graph: OpGraph, ctx: TransformContext) -> None:
        """Write all transformed operators with detailed annotations."""
        ws = wb.create_sheet("Transformed Operators")

        columns = [
            ("Node ID", 12),
            ("Op Type", 25),
            ("Category", 12),
            ("Scope", 45),
            ("Layer", 7),
            ("Component", 18),
            ("Parallelism Strategy", 18),
            ("Collective Op", 15),
            ("Group Size", 10),
            ("Role", 10),
            ("Pipeline Stage", 14),
            ("Stream Type", 15),
            ("Stream ID", 10),
            ("Input Shapes", 50),
            ("Output Shapes", 50),
            ("Input Dtypes", 20),
            ("Output Dtypes", 20),
            ("FLOPs", 12),
            ("Compute (µs)", 12),
            ("Memory (µs)", 12),
            ("Total Latency (µs)", 14),
            ("Bound", 10),
            ("Arithmetic Intensity", 16),
            ("Annotations", 60),
        ]

        self._write_header(ws, columns)

        # Infer layer-to-stage mapping for pipeline parallelism
        layer_to_stage = {}
        if ctx.parallel.pp > 1:
            layers = set(n.layer for n in graph.nodes.values() if n.layer)
            sorted_layers = sorted(layers, key=lambda x: int(x) if x.isdigit() else 0)
            for i, layer in enumerate(sorted_layers):
                layer_to_stage[layer] = i % ctx.parallel.pp

        for row_idx, node in enumerate(graph.topo_sort(), 2):
            parallelism = get_parallelism_info(node, ctx.parallel)

            input_shapes = ", ".join(str(t.shape) for t in node.inputs)
            output_shapes = ", ".join(str(t.shape) for t in node.outputs)
            input_dtypes = ", ".join(str(t.dtype) for t in node.inputs)
            output_dtypes = ", ".join(str(t.dtype) for t in node.outputs)

            # Build annotations string
            annotations_list = []
            for key, val in node.annotations.items():
                if key not in ("stream_id", "stream_type", "flops", "compute_us",
                               "memory_us", "latency_us", "arithmetic_intensity", "bound",
                               "read_bytes", "write_bytes"):
                    if isinstance(val, dict):
                        annotations_list.append(f"{key}={str(val)[:30]}")
                    else:
                        annotations_list.append(f"{key}={val}")
            annotations_str = "; ".join(annotations_list) if annotations_list else ""

            values = [
                node.id,
                node.op_type,
                node.category,
                node.scope,
                node.layer or "",
                node.component,
                parallelism["strategy"],
                parallelism.get("collective", ""),
                parallelism.get("group_size", ""),
                parallelism.get("role", ""),
                infer_pipeline_stage(node, layer_to_stage),
                node.annotations.get("stream_type", ""),
                node.annotations.get("stream_id", ""),
                input_shapes,
                output_shapes,
                input_dtypes,
                output_dtypes,
                node.annotations.get("flops", ""),
                round(node.annotations.get("compute_us", 0), 3) if node.annotations.get("compute_us") else "",
                round(node.annotations.get("memory_us", 0), 3) if node.annotations.get("memory_us") else "",
                round(node.annotations.get("latency_us", 0), 3) if node.annotations.get("latency_us") else "",
                node.annotations.get("bound", ""),
                round(node.annotations.get("arithmetic_intensity", 0), 2) if node.annotations.get("arithmetic_intensity") else "",
                annotations_str,
            ]

            # Choose fill color based on category
            if node.is_comm:
                fill = self._comm_fill
            elif node.category == "memory":
                fill = self._memory_fill
            else:
                fill = self._compute_fill

            self._write_row(ws, row_idx, values, fill)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(graph.nodes) + 1}"
        ws.freeze_panes = "A2"

    def _write_communication_sheet(self, wb: openpyxl.Workbook,
                                   graph: OpGraph, ctx: TransformContext) -> None:
        """Write communication operators with collective details."""
        ws = wb.create_sheet("Communication Ops")

        comm_nodes = [n for n in graph.nodes.values() if n.is_comm]
        if not comm_nodes:
            ws.append(["No communication operators found"])
            return

        columns = [
            ("Node ID", 12),
            ("Collective Op", 15),
            ("Role", 10),
            ("Group Size", 10),
            ("Scope", 45),
            ("Layer", 7),
            ("Stream Type", 15),
            ("Stream ID", 10),
            ("Input Shapes", 50),
            ("Output Shapes", 50),
            ("Inserted By", 15),
            ("Data Volume (bytes)", 18),
        ]

        self._write_header(ws, columns)

        for row_idx, node in enumerate(comm_nodes, 2):
            collective = node.attrs.get("collective", "")
            group_size = node.attrs.get("group_size", "")
            role = node.attrs.get("role", "")

            # Estimate data volume: sum of output tensor sizes
            data_volume = sum(t.mem_bytes for t in node.outputs)

            input_shapes = ", ".join(str(t.shape) for t in node.inputs)
            output_shapes = ", ".join(str(t.shape) for t in node.outputs)

            values = [
                node.id,
                collective,
                role,
                group_size,
                node.scope,
                node.layer or "",
                node.annotations.get("stream_type", ""),
                node.annotations.get("stream_id", ""),
                input_shapes,
                output_shapes,
                node.annotations.get("inserted_by", ""),
                data_volume,
            ]

            self._write_row(ws, row_idx, values, self._comm_fill)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(comm_nodes) + 1}"
        ws.freeze_panes = "A2"
        logger.info(f"Found {len(comm_nodes)} communication operators")

    def _write_parallelism_summary_sheet(self, wb: openpyxl.Workbook,
                                        graph: OpGraph, ctx: TransformContext) -> None:
        """Write summary of parallelism distribution across layers."""
        ws = wb.create_sheet("Parallelism Summary")

        # Group by layer
        layer_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
            "compute_count": 0, "comm_count": 0, "memory_count": 0,
            "comm_ops": set(), "inserted_by": []
        })

        for node in graph.nodes.values():
            layer = node.layer or "non-layer"

            if node.is_comm:
                layer_stats[layer]["comm_count"] += 1
                collective = node.attrs.get("collective", "")
                layer_stats[layer]["comm_ops"].add(collective)
                inserted_by = node.annotations.get("inserted_by", "")
                if inserted_by:
                    layer_stats[layer]["inserted_by"].append(inserted_by)
            elif node.category == "memory":
                layer_stats[layer]["memory_count"] += 1
            else:
                layer_stats[layer]["compute_count"] += 1

        ws.append(["Layer", "Compute Ops", "Memory Ops", "Comm Ops", "Comm Types", "Inserted By", "Parallelism Strategy"])
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            ws[f"{col}1"].font = self._header_font
            ws[f"{col}1"].fill = self._header_fill

        for layer in sorted(layer_stats.keys(), key=lambda x: (x == "non-layer", x)):
            stats = layer_stats[layer]
            comm_types = ", ".join(sorted(stats["comm_ops"])) if stats["comm_ops"] else ""
            inserted_by = ", ".join(sorted(set(stats["inserted_by"]))) if stats["inserted_by"] else ""

            # Infer parallelism strategy for this layer
            parallel_strat = []
            if stats["comm_count"] > 0:
                if "all_reduce" in stats["comm_ops"]:
                    parallel_strat.append("TP")
                if "all_to_all" in stats["comm_ops"]:
                    parallel_strat.append("EP")
            parallel_strat_str = "/".join(parallel_strat) if parallel_strat else "no-parallel"

            ws.append([
                layer,
                stats["compute_count"],
                stats["memory_count"],
                stats["comm_count"],
                comm_types,
                inserted_by,
                parallel_strat_str,
            ])

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20

    def _write_stream_assignment_sheet(self, wb: openpyxl.Workbook,
                                      graph: OpGraph, ctx: TransformContext) -> None:
        """Write stream assignment details."""
        ws = wb.create_sheet("Stream Assignment")

        # Collect stream statistics
        stream_assignment: Dict[int, List[str]] = defaultdict(list)
        stream_types: Dict[int, str] = {}

        for node in graph.nodes.values():
            stream_id = node.annotations.get("stream_id")
            stream_type = node.annotations.get("stream_type")
            if stream_id is not None:
                stream_assignment[stream_id].append(node.id)
                stream_types[stream_id] = stream_type or "unknown"

        ws.append(["Stream ID", "Stream Type", "Assigned Nodes", "Node Count"])
        for col in ("A", "B", "C", "D"):
            ws[f"{col}1"].font = self._header_font
            ws[f"{col}1"].fill = self._header_fill

        for stream_id in sorted(stream_assignment.keys()):
            node_ids = stream_assignment[stream_id]
            stream_type = stream_types[stream_id]

            ws.append([
                stream_id,
                stream_type,
                "; ".join(node_ids),
                len(node_ids),
            ])

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 12

    def _write_header(self, ws, columns: List[tuple[str, int]]) -> None:
        """Write header row with styling."""
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_row(self, ws, row_idx: int, values: List[Any],
                   fill: Optional[PatternFill] = None) -> None:
        """Write data row with optional fill."""
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = self._thin_border
            if fill:
                cell.fill = fill
            if col_idx > 10:  # Wrap text for longer content
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def export_transformed_graph(graph: OpGraph, ctx: TransformContext,
                            output_dir: Path) -> Dict[str, Path]:
    """Export transformed graph to Excel, JSON, and optionally ONNX.

    Parameters
    ----------
    graph : OpGraph
        The transformed computation graph
    ctx : TransformContext
        Transformation context with parallel config and stream config
    output_dir : Path
        Output directory for exported files

    Returns
    -------
    dict[str, Path]
        Paths to generated files: {format: path}
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Determine output filename
    base_name = graph.name.replace("/", "_").replace(":", "_")

    # Excel export
    excel_path = output_dir / f"{base_name}_transformed_ops.xlsx"
    writer = TransformedGraphExcelWriter()
    writer.write(graph, ctx, excel_path)

    # JSON export (simplified)
    json_path = output_dir / f"{base_name}_transformed_graph.json"
    _export_json(graph, ctx, json_path)

    return {
        "excel": excel_path,
        "json": json_path,
    }


def _export_json(graph: OpGraph, ctx: TransformContext, output_path: Path) -> None:
    """Export transformed graph to JSON format."""
    data = {
        "graph": {
            "name": graph.name,
            "phase": graph.phase,
            "num_nodes": len(graph.nodes),
            "num_edges": len(graph.edges),
        },
        "parallelism": {
            "strategy": ctx.parallel.describe(),
            "tp": ctx.parallel.tp,
            "ep": ctx.parallel.ep,
            "pp": ctx.parallel.pp,
            "dp": ctx.parallel.dp,
            "sp": ctx.parallel.sp,
        },
        "stream_config": {
            "compute_streams": ctx.stream_config.num_compute_streams,
            "comm_streams": ctx.stream_config.num_comm_streams,
        },
        "nodes": [],
        "edges": [],
    }

    # Convert nodes
    for node in graph.topo_sort():
        node_data = {
            "id": node.id,
            "op_type": node.op_type,
            "category": node.category,
            "scope": node.scope,
            "layer": node.layer,
            "attrs": node.attrs,
            "annotations": node.annotations,
            "input_shapes": [list(t.shape) for t in node.inputs],
            "output_shapes": [list(t.shape) for t in node.outputs],
        }
        data["nodes"].append(node_data)

    # Convert edges
    for edge in graph.edges:
        edge_data = {
            "src": edge.src,
            "dst": edge.dst,
            "src_idx": edge.src_idx,
            "dst_idx": edge.dst_idx,
        }
        data["edges"].append(edge_data)

    output_path.write_text(json.dumps(data, indent=2, default=str))
    logger.info(f"Exported transformed graph JSON to {output_path}")
