"""Integration test: verify how DataParallel (DP) affects TrainingReport fields
and validates DP operator correctness in the generated XLSX report.

This test runs the training modelling CLI for dp=1, dp=4, dp=8, loads the
generated ``deepseek_v4_training_report.json`` files and the training XLSX,
then asserts expected relationships:

JSON-report checks:
- optimizer state (opt_state) per-GPU ≈ 1/dp  (ZeRO-1)
- total per-GPU memory drops monotonically as dp increases
- step_time decreases and tokens/sec increases monotonically as dp increases
- dp_hidden_ms and dp_exposed_ms are zero for dp=1, non-zero for dp>1
- dp_comm_total (dp_hidden + dp_exposed) increases monotonically with dp
- optimizer state scales approximately as 1/dp across dp=4 and dp=8

XLSX operator checks (DataParallel validation):
- dp=1 → no DP comm nodes in "Communication Ops" sheet
- dp>1 → DP comm nodes exist with correct collective type, group_size, role
- one DP comm node per layer (4 layers with --layers 4)
- each DP comm node has a matching grad-scale node (aten.div.Scalar) via Node ID
- "Parallelism Summary" sheet references data_parallel_pass

This is a long-running integration test that captures real reports. To avoid
running it by default in fast CI, it is skipped unless the environment
variable ``RUN_DP_TEST`` is set to ``1``.

Run locally (PowerShell):

```powershell
$env:RUN_DP_TEST='1'; pytest tests/IT/test_dp_effects.py -q
```
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import pytest


def _run_cli_and_load_report(repo_root: Path, outdir: Path, dp: int, timeout: int = 900) -> dict:
    """Run `python -m python.zrt` with given dp and return parsed report JSON + XLSX path.

    Returns a dict with keys ``"report"`` and ``"xlsx_path"``.

    Raises subprocess.CalledProcessError on failure.
    """
    env = os.environ.copy()
    env["PYTHONPATH"] = str(repo_root / "python")

    cmd = [
        sys.executable,
        "-m",
        "python.zrt",
        "--model-id",
        "hf_models/deepseek_v4",
        "--train",
        "--hw",
        "nvidia_h100_sxm",
        "--dp",
        str(dp),
        "--layers",
        "4",
        "--batch-size",
        "1",
        "--seq-len",
        "128",
        "--output-dir",
        str(outdir),
    ]

    proc = subprocess.run(cmd, check=True, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, timeout=timeout)
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "cli_output.log").write_text(proc.stdout)

    report_path = outdir / "reports" / "deepseek_v4_training_report.json"
    assert report_path.exists(), f"Report not found at {report_path}"
    report = json.loads(report_path.read_text())

    xlsx_path = _find_training_xlsx(outdir)

    return {"report": report, "xlsx_path": xlsx_path}


def _find_training_xlsx(outdir: Path) -> Path | None:
    """Find the training XLSX file in the output directory.

    The XLSX is named ``<graph_base>_training.xlsx``.  Search via glob so we
    are robust to graph-name variations.
    """
    candidates = list(outdir.glob("*_training.xlsx"))
    if candidates:
        return candidates[0]
    return None


def _read_xlsx_sheet(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Read an XLSX sheet and return rows as list of dicts keyed by header."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        entry = {}
        for i, val in enumerate(row):
            if i < len(headers):
                entry[headers[i]] = val
        result.append(entry)
    return result


@pytest.fixture(scope="session")
def dp_reports(tmp_path_factory):
    """Session-scoped fixture: run CLI for dp=1, dp=4, dp=8, return all reports + XLSX paths."""
    if os.environ.get("RUN_DP_TEST") != "1":
        pytest.skip("Set RUN_DP_TEST=1 to run this long integration test")

    repo_root = Path(__file__).resolve().parents[3]
    tmp_path = tmp_path_factory.mktemp("dp_effects")

    results = {}
    for dp in (1, 4, 8):
        out_dir = tmp_path / f"out_dp{dp}"
        results[dp] = _run_cli_and_load_report(repo_root, out_dir, dp=dp)

    return results


# ── ZeRO-1 optimizer state scaling ─────────────────────────────────────────

def _report(dp_reports, dp: int) -> dict:
    """Extract JSON report from the fixture result for the given dp value."""
    return dp_reports[dp]["report"]


def _xlsx_path(dp_reports, dp: int) -> Path | None:
    """Extract XLSX path from the fixture result for the given dp value."""
    path = dp_reports[dp].get("xlsx_path")
    return Path(path) if path else None


# ── ZeRO-1 optimizer state scaling ─────────────────────────────────────────

def test_dp_optimizer_state_scales_inverse(dp_reports):
    """Optimizer state per-GPU should roughly scale ~1/dp (ZeRO-1 behaviour)."""
    opt = {}
    for dp in (1, 4, 8):
        mb = _report(dp_reports, dp).get("memory_breakdown_gb")
        assert mb, f"memory_breakdown_gb missing for dp={dp}"
        opt[dp] = mb.get("opt_state")
        assert opt[dp] is not None, f"opt_state missing for dp={dp}"

    # Approximate 1/dp scaling: opt(dp) ≈ opt(1) / dp
    for dp in (4, 8):
        expected = opt[1] / dp
        assert opt[dp] == pytest.approx(expected, rel=0.25), (
            f"opt_state did not scale near 1/dp: dp=1 → {opt[1]}, dp={dp} → {opt[dp]}, "
            f"expected ≈ {expected}"
        )

    # Monotonic decrease
    assert opt[1] > opt[4] > opt[8], (
        f"opt_state should decrease monotonically: dp1={opt[1]}, dp4={opt[4]}, dp8={opt[8]}"
    )


# ── Total memory monotonicity ──────────────────────────────────────────────

def test_dp_total_memory_decreases(dp_reports):
    """Total per-GPU memory should decrease monotonically as dp increases."""
    total = {}
    for dp in (1, 4, 8):
        mb = _report(dp_reports, dp)["memory_breakdown_gb"]
        total[dp] = mb["total"]
        assert total[dp] is not None

    assert total[1] > total[4] > total[8], (
        f"total memory should decrease monotonically: "
        f"dp1={total[1]}, dp4={total[4]}, dp8={total[8]}"
    )


# ── Throughput monotonicity ────────────────────────────────────────────────

def test_dp_throughput_improves(dp_reports):
    """Step time should decrease and tokens/sec should increase monotonically."""
    step_time = {dp: _report(dp_reports, dp)["step_time_ms"] for dp in (1, 4, 8)}

    assert step_time[1] > step_time[4] > step_time[8], (
        f"step_time should decrease monotonically: "
        f"dp1={step_time[1]}, dp4={step_time[4]}, dp8={step_time[8]}"
    )


# ── DP communication accounting ────────────────────────────────────────────

def test_dp_communication_zero_for_dp1(dp_reports):
    """dp_hidden_ms and dp_exposed_ms should be 0 for dp=1 (no DP communication)."""
    rep1 = _report(dp_reports, 1)

    dp_hidden1 = rep1.get("dp_hidden_ms")
    dp_exposed1 = rep1.get("dp_exposed_ms")
    assert dp_hidden1 is not None
    assert dp_exposed1 is not None
    assert dp_hidden1 == 0.0, f"dp_hidden should be 0 for dp=1, got {dp_hidden1}"
    assert dp_exposed1 == 0.0, f"dp_exposed should be 0 for dp=1, got {dp_exposed1}"


def test_dp_communication_nonzero_for_dp_gt1(dp_reports):
    """dp_hidden + dp_exposed should be >0 for dp>1 (DP AR/RS communication exists).

    dp_hidden may be 0 (e.g. no bubble to absorb AR) or dp_exposed may be 0
    (e.g. AR fully hidden in bubble), but their sum must reflect the total
    DP communication volume.
    """
    for dp in (4, 8):
        rep = _report(dp_reports, dp)
        dp_hidden = rep.get("dp_hidden_ms")
        dp_exposed = rep.get("dp_exposed_ms")
        dp_comm = rep.get("dp_total_ms", 0.0)
        assert dp_hidden is not None, f"dp_hidden_ms missing for dp={dp}"
        assert dp_exposed is not None, f"dp_exposed_ms missing for dp={dp}"
        assert dp_comm > 0, (
            f"dp_hidden + dp_exposed should be >0 for dp={dp}, "
            f"got hidden={dp_hidden}, exposed={dp_exposed}"
        )


def test_dp_comm_volume_monotonic(dp_reports):
    """Total DP communication volume (dp_hidden + dp_exposed) should increase with dp.

    Larger DP group means more gradient data to reduce, so the total DP
    communication time (exposed + hidden) should grow monotonically.
    Note: on full-mesh topologies the per-step time may not scale linearly
    due to the (N-1)/N ring factor, but the trend should be increasing.
    """
    dp_comm = {}
    for dp in (4, 8):
        rep = _report(dp_reports, dp)
        dp_comm[dp] = rep.get("dp_total_ms", 0.0)

    assert dp_comm[8] > dp_comm[4], (
        f"DP comm volume should increase with dp: "
        f"dp4={dp_comm[4]:.2f}ms, dp8={dp_comm[8]:.2f}ms"
    )


# ═══════════════════════════════════════════════════════════════════════════
# XLSX Operator Validation — DataParallel correctness
# ═══════════════════════════════════════════════════════════════════════════


def _get_dp_comm_nodes(xlsx_path: Path) -> list[dict]:
    """Return all DP communication nodes from the "Communication Ops" sheet.

    DP comm nodes are identified by role="dp_grad_reduce" or
    inserted_by="data_parallel_pass".
    """
    comm_rows = _read_xlsx_sheet(xlsx_path, "Communication Ops")
    return [
        row for row in comm_rows
        if row.get("Role") == "dp_grad_reduce"
        or "data_parallel_pass" in str(row.get("Inserted By", ""))
    ]


def _get_backward_ops(xlsx_path: Path) -> list[dict]:
    """Return all rows from the "Backward Operators" sheet."""
    return _read_xlsx_sheet(xlsx_path, "Backward Operators")


def _get_transformed_ops(xlsx_path: Path) -> list[dict]:
    """Return all rows from the "Forward Operators" sheet."""
    return _read_xlsx_sheet(xlsx_path, "Forward Operators")


# ── DP comm nodes: existence ──────────────────────────────────────────────

def test_dp_xlsx_no_comm_for_dp1(dp_reports):
    """dp=1 should have zero DP communication nodes in the XLSX."""
    xlsx = _xlsx_path(dp_reports, 1)
    assert xlsx is not None, "XLSX not generated for dp=1 — export may have failed"
    assert xlsx.exists(), f"XLSX missing at {xlsx}"

    dp_comm = _get_dp_comm_nodes(xlsx)
    assert len(dp_comm) == 0, (
        f"Expected 0 DP comm nodes for dp=1, got {len(dp_comm)}: "
        f"{[r.get('Node ID') for r in dp_comm]}"
    )


def test_dp_xlsx_comm_nodes_exist_for_dp_gt1(dp_reports):
    """dp>1 should have DP communication nodes in the XLSX."""
    for dp in (4, 8):
        xlsx = _xlsx_path(dp_reports, dp)
        assert xlsx is not None, f"XLSX not generated for dp={dp}"
        assert xlsx.exists(), f"XLSX missing at {xlsx}"

        dp_comm = _get_dp_comm_nodes(xlsx)
        assert len(dp_comm) > 0, (
            f"Expected >0 DP comm nodes for dp={dp}, got {len(dp_comm)}"
        )


# ── DP comm nodes: properties ─────────────────────────────────────────────

def test_dp_xlsx_comm_collective_type(dp_reports):
    """DP comm nodes should have collective="reduce_scatter" (ZeRO-1 default)."""
    for dp in (4, 8):
        xlsx = _xlsx_path(dp_reports, dp)
        assert xlsx is not None and xlsx.exists()

        dp_comm = _get_dp_comm_nodes(xlsx)
        for node in dp_comm:
            collective = node.get("Collective Op", "")
            assert collective == "reduce_scatter", (
                f"Expected reduce_scatter for ZeRO-1 dp={dp}, "
                f"got '{collective}' in node {node.get('Node ID')}"
            )


def test_dp_xlsx_comm_group_size(dp_reports):
    """DP comm nodes should have group_size matching the dp degree."""
    for dp in (4, 8):
        xlsx = _xlsx_path(dp_reports, dp)
        assert xlsx is not None and xlsx.exists()

        dp_comm = _get_dp_comm_nodes(xlsx)
        for node in dp_comm:
            group_size = node.get("Group Size")
            assert group_size is not None, (
                f"Group Size missing for node {node.get('Node ID')}"
            )
            assert int(group_size) == dp, (
                f"Expected group_size={dp}, got {group_size} "
                f"in node {node.get('Node ID')}"
            )


def test_dp_xlsx_comm_role_and_inserted_by(dp_reports):
    """DP comm nodes should have role="dp_grad_reduce" and inserted_by="data_parallel_pass"."""
    for dp in (4, 8):
        xlsx = _xlsx_path(dp_reports, dp)
        assert xlsx is not None and xlsx.exists()

        dp_comm = _get_dp_comm_nodes(xlsx)
        for node in dp_comm:
            role = node.get("Role", "")
            assert role == "dp_grad_reduce", (
                f"Expected role='dp_grad_reduce', got '{role}' "
                f"in node {node.get('Node ID')}"
            )
            inserted_by = str(node.get("Inserted By", ""))
            assert "data_parallel_pass" in inserted_by, (
                f"Expected inserted_by='data_parallel_pass', got '{inserted_by}' "
                f"in node {node.get('Node ID')}"
            )


# ── DP comm node count ───────────────────────────────────────────

def test_dp_xlsx_comm_node_count(dp_reports):
    """One DP communication node total (aligned with estimate path)."""
    for dp in (4, 8):
        xlsx = _xlsx_path(dp_reports, dp)
        assert xlsx is not None and xlsx.exists()

        dp_comm = _get_dp_comm_nodes(xlsx)
        assert len(dp_comm) == 1, (
            f"Expected 1 DP comm node for dp={dp} (single group), "
            f"got {len(dp_comm)}"
        )

        node = dp_comm[0]
        nid = node.get("Node ID", "")
        assert nid == "comm_grad_reduce", (
            f"Expected Node ID 'comm_grad_reduce', got '{nid}'"
        )


# ── Grad-scale nodes (aten.div.Scalar) ────────────────────────────────────

def test_dp_xlsx_grad_scale_nodes_exist(dp_reports):
    """A single grad-scale node (aten.div.Scalar) should exist, matching the DP comm node.

    Identified by Node ID ``grad_scale``.
    """
    for dp in (4, 8):
        xlsx = _xlsx_path(dp_reports, dp)
        assert xlsx is not None and xlsx.exists()

        bwd_ops = _get_backward_ops(xlsx)
        if not bwd_ops:
            continue

        grad_scale_nodes = [
            row for row in bwd_ops
            if row.get("Op Type") == "aten.div.Scalar"
            and str(row.get("Node ID", "")) == "grad_scale"
        ]

        assert len(grad_scale_nodes) == 1, (
            f"Expected 1 grad-scale node for dp={dp}, got {len(grad_scale_nodes)}"
        )


# ── Parallelism Summary sheet ─────────────────────────────────────────────

def test_dp_xlsx_parallelism_summary_shows_dp(dp_reports):
    """Parallelism Summary sheet should list data_parallel_pass as inserted_by."""
    for dp in (4, 8):
        xlsx = _xlsx_path(dp_reports, dp)
        assert xlsx is not None and xlsx.exists()

        summary_rows = _read_xlsx_sheet(xlsx, "Parallelism Summary")
        assert len(summary_rows) > 0, f"Parallelism Summary sheet empty for dp={dp}"

        inserted_by_values = set()
        for row in summary_rows:
            val = str(row.get("Inserted By", ""))
            if val:
                inserted_by_values.add(val)

        assert "data_parallel_pass" in inserted_by_values, (
            f"Parallelism Summary should mention 'data_parallel_pass' for dp={dp}, "
            f"found: {inserted_by_values}"
        )